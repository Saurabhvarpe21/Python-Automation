import os
import pandas as pd
from openpyxl import load_workbook
f=pd.DataFrame(columns=['commodity_name','Not_Start "S" And 13 Char','G06 Folder is Empty'])

def validate_source_folder(folderName):
    if(folderName.startswith("S")):
        if(len(folderName)==13):
            return "valid"
        else:
            return "Folder name not 13 characters"
    else:
        return "Folder name not starting with S."

def massage_directory_path(rawpath):
    print(rawpath)
    return rawpath.replace('\\', '\\\\')
        
def RT_NA_to_CUSTOM_IMAGE_fn():
    t=1
    print("Please enter the below details")
    source_folder = massage_directory_path(input("Enter the source directory: "));
    # get all the first level folders
    allFilesAndFolders =os.listdir(source_folder)
    
    for i in allFilesAndFolders:
        fullPath = source_folder + "\\" + i;
#         #Access first level folder
        if os.path.isdir(fullPath):
            #Source folder validation for S and 13 charecters
            validationOutput = validate_source_folder(i)
            if(validationOutput == "valid"):
                g06fullPath = fullPath + "\\" + "G06 TABLES"
#                 Check for G06 TABLES
                if os.path.isdir(g06fullPath):
                    allG06FilesAndFolders =os.listdir(g06fullPath)
#                     print(allG06FilesAndFolders)
                    if allG06FilesAndFolders:
                        for j in allG06FilesAndFolders:
                            fp = g06fullPath + "\\" 
                            currentObj = fp + j;
                            if(os.path.isfile(currentObj)):
    #                             #Form new file name
                                oldFileFullPath = fp+j;
                                print("oldFileFullPath: ",oldFileFullPath );
                                if(j.endswith(".xlsx")) and (j.startswith("S")):
                                    df=pd.read_excel(oldFileFullPath)
                                    wb = load_workbook(oldFileFullPath)
                                    ws = wb._sheets[0]
                                    cl = df.columns
                                    le = df.shape[0]
                                    for r in range (0,le):
                                        if str(df[cl[8]][r]) == 'nan' and str(df[cl[9]][r]) == 'WIRE' and str(df[cl[11]][r])!= 'nan':
                                            ws[f'I{r+2}'].value = 'CUSTOM IMAGE'
                                        else:
                                            pass
                                    wb.save(oldFileFullPath)

                                else:
                                    os.remove(oldFileFullPath)                        
                    else:
                        print("empty g06 folder")
                        f.loc[t,'commodity_name']=i
                        f.loc[t,'G06 Folder is Empty']="✓"
                        t+=1
                else:
                    f.loc[t,'commodity_name']=i
                    f.loc[t,'G06 TABLES folder not found']="✓"
                    t+=1
                    print(i + " : " + validationOutput)
            else:
                f.loc[t,'commodity_name']=i
                f.loc[t,'Not_Start "S" And 13 Char']="✓"
                t+=1

        f.to_excel('RT_NA_TO_CUSTOM_IMAGE.xlsx',index=False)
              
